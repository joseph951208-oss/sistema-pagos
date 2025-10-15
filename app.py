from flask import Flask, render_template, request, redirect, url_for, session, flash
from openpyxl import load_workbook
from werkzeug.security import check_password_hash
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime
from sqlalchemy import func
import os
import json

app = Flask(__name__)
app.secret_key = 'clave-secreta-muy-segura'

# URL de PostgreSQL 
DATABASE_URL = os.environ.get(
    "DATABASE_URL",
    "postgresql://sistema_pagos_ohrc_user:LsMj7GgLXTPIW2C7rTvE3kfCjQz4j9OW@dpg-d3jh63t6ubrc73cr05ag-a.oregon-postgres.render.com/sistema_pagos_ohrc"
)
app.config['SQLALCHEMY_DATABASE_URI'] = DATABASE_URL
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

# -------------------- Modelos --------------------
class Cliente(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    cedula = db.Column(db.String(20), unique=True, nullable=False)
    nombres = db.Column(db.String(100), nullable=False)
    direccion = db.Column(db.String(200))
    fecha_instalacion = db.Column(db.Date)
    ip = db.Column(db.String(50))
    pagos = db.relationship('Pago', backref='cliente', lazy=True)

class Pago(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    cliente_id = db.Column(db.Integer, db.ForeignKey('cliente.id'), nullable=False)
    documento = db.Column(db.String(50))
    fecha_pago = db.Column(db.Date, nullable=False)
    forma_pago = db.Column(db.String(50), nullable=False)

# -------------------- Configuración de uploads --------------------
UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# -------------------- Función para cargar documentos --------------------
def cargar_documentos():
    documentos = []
    try:
        movimientos_path = os.path.join(app.config['UPLOAD_FOLDER'], 'movimientos.xlsx')
        if not os.path.exists(movimientos_path):
            return []

        wb = load_workbook(movimientos_path, data_only=True)
        ws = wb.active
        header_row = 6  # Fila donde está "Nro. Documento"
        col_idx = None

        # Buscar columna sin importar mayúsculas, puntos o espacios
        for cell in ws[header_row]:
            if cell.value and "NRO" in str(cell.value).upper() and "DOCUMENTO" in str(cell.value).upper():
                col_idx = cell.column
                break

        if not col_idx:
            return []

        merged_ranges = ws.merged_cells.ranges
        for row in range(header_row + 1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            val = cell.value

            # Si la celda está vacía, verificar si pertenece a un rango combinado
            if not val:
                for mrange in merged_ranges:
                    if cell.coordinate in mrange:
                        val = ws.cell(mrange.min_row, mrange.min_col).value
                        break

            # Validar que el valor no esté vacío ni sea NaN
            if val and str(val).strip() not in ["", "NaN", "nan"]:
                val_str = str(val).strip()
                if val_str.endswith(".0"):
                    val_str = val_str[:-2]
                documentos.append(val_str)

    except Exception as e:
        print(f"❌ Error al cargar movimientos.xlsx: {e}")
        return []

    return documentos

# -------------------- Documentos precargados --------------------
DOCUMENTOS_CARGADOS = []
with app.app_context():
    DOCUMENTOS_CARGADOS = cargar_documentos()

# -------------------- Cargar usuarios desde JSON --------------------
with open("usuarios.json", "r") as f:
    USERS = json.load(f)

# -------------------- Rutas --------------------
@app.route('/', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')

        user = USERS.get(username)

        if user and check_password_hash(user['password_hash'], password):
            session['username'] = username
            session['role'] = user['role']

            if user['role'] == 'admin':
                return redirect(url_for('upload_files'))
            elif user['role'] == 'reader':
                return redirect(url_for('consulta'))

        flash("Usuario o contraseña incorrectos", "danger")

    return render_template('login.html')

# -------------------- Subir archivos --------------------
@app.route('/subir', methods=['GET', 'POST'])
def upload_files():
    if 'username' not in session or session.get('role') != 'admin':
        flash('No tienes permisos para acceder a esta página.', 'danger')
        return redirect(url_for('login'))

    global DOCUMENTOS_CARGADOS

    if request.method == 'POST':
        movimientos_file = request.files.get('movimientos')

        if not movimientos_file or movimientos_file.filename == '':
            flash('Debes seleccionar un archivo de movimientos.', 'danger')
            return redirect(url_for('upload_files'))

        movimientos_path = os.path.join(app.config['UPLOAD_FOLDER'], 'movimientos.xlsx')
        movimientos_file.save(movimientos_path)

        # 🔹 Cargar documentos y eliminar duplicados
        DOCUMENTOS_CARGADOS = sorted(list(set(cargar_documentos())))

        flash('Archivo de movimientos cargado exitosamente ✅', 'success')
        return redirect(url_for('upload_files'))

    return render_template('index.html')

# -------------------- Registro de Pagos --------------------
@app.route('/registro_pago', methods=['GET', 'POST'])
def registro_pago():
    if 'username' not in session:
        return redirect(url_for('login'))

    clientes_db = Cliente.query.all()
    clientes = [c.nombres for c in clientes_db]

    documentos = sorted(list(set(DOCUMENTOS_CARGADOS)))  # 🔹 Únicos y ordenados

    if request.method == 'POST':
        cliente_nombre = request.form.get('cliente', '').strip()
        fecha_pago = request.form.get('fecha_pago')
        forma_pago = request.form.get('forma_pago')
        doc_num = request.form.get('doc_num', '').strip()

        cliente = Cliente.query.filter(func.lower(func.trim(Cliente.nombres)) == cliente_nombre.lower().strip()).first()
        if not cliente:
            flash("Cliente no válido o no registrado.", "danger")
            return redirect(url_for('registro_pago'))

        if forma_pago.upper() == "TRANSFERENCIA":
            if not doc_num:
                flash("Debe ingresar un número de documento para transferencia.", "danger")
                return redirect(url_for('registro_pago'))
            if doc_num not in documentos:
                flash(f"El número de documento {doc_num} no se encuentra en movimientos.xlsx.", "danger")
                return redirect(url_for('registro_pago'))

        if forma_pago.upper() == "OTROS BANCOS" and not doc_num:
            doc_num = ""

        pago_existente = Pago.query.filter_by(documento=doc_num).first()
        if pago_existente and doc_num != "":
            flash(f"El documento {doc_num} ya fue registrado.", "warning")
            return redirect(url_for('registro_pago'))

        try:
            nuevo_pago = Pago(
                cliente_id=cliente.id,
                documento=doc_num or "",
                fecha_pago=datetime.strptime(fecha_pago, "%Y-%m-%d"),
                forma_pago=forma_pago
            )
            db.session.add(nuevo_pago)
            db.session.commit()
            flash("✅ Pago registrado exitosamente.", "success")
        except Exception as e:
            db.session.rollback()
            flash(f"❌ Error al registrar el pago: {str(e)}", "danger")

        return redirect(url_for('registro_pago'))

    return render_template('registro_pago.html', clientes=clientes, documentos=documentos)

# -------------------- Consulta --------------------
@app.route('/consulta', methods=['GET'])
def consulta():
    if 'username' not in session:
        return redirect(url_for('login'))

    cedula = request.args.get('cedula', '').strip()
    cliente_nombre = request.args.get('cliente', '').strip()
    documento = request.args.get('documento', '').strip()

    query = db.session.query(
        Cliente.cedula.label('Cedula'),
        Cliente.nombres.label('Cliente'),
        Pago.documento.label('Documento'),
        Pago.fecha_pago.label('Fecha'),
        Pago.forma_pago.label('Forma')
    ).join(Pago, Pago.cliente_id == Cliente.id)

    if cedula:
        query = query.filter(Cliente.cedula.ilike(f"%{cedula}%"))
    if cliente_nombre:
        query = query.filter(Cliente.nombres.ilike(f"%{cliente_nombre}%"))
    if documento:
        query = query.filter(Pago.documento.ilike(f"%{documento}%"))

    pagos = query.order_by(Pago.fecha_pago.desc()).all()

    pagos_lista = [dict(
        Cedula=p.Cedula,
        Cliente=p.Cliente,
        Documento=p.Documento,
        Fecha=p.Fecha.strftime("%Y-%m-%d"),
        Forma=p.Forma
    ) for p in pagos]

    clientes = [c.nombres for c in Cliente.query.order_by(Cliente.nombres).all()]
    return render_template('consulta.html', pagos=pagos_lista, clientes=clientes)

# -------------------- Registro de Clientes --------------------
@app.route('/registro_cliente', methods=['GET', 'POST'])
def registro_cliente():
    if 'username' not in session:
        return redirect(url_for('login'))

    mensaje = None

    if request.method == 'POST':
        accion = request.form.get('accion')

        if accion == "agregar":
            cedula = request.form.get('cedula').strip()
            nombre = request.form.get('nombre').strip()
            direccion = request.form.get('direccion').strip()
            fecha_instalacion = request.form.get('fecha_instalacion')
            ip = request.form.get('ip').strip()

            if cedula and nombre:
                if Cliente.query.filter_by(cedula=cedula).first():
                    mensaje = f"❌ Cliente con cédula {cedula} ya existe."
                else:
                    try:
                        nuevo_cliente = Cliente(
                            cedula=cedula,
                            nombres=nombre,
                            direccion=direccion,
                            fecha_instalacion=datetime.strptime(fecha_instalacion, "%Y-%m-%d") if fecha_instalacion else None,
                            ip=ip
                        )
                        db.session.add(nuevo_cliente)
                        db.session.commit()
                        mensaje = "✅ Cliente agregado con éxito."
                    except Exception as e:
                        db.session.rollback()
                        mensaje = f"❌ Error al agregar cliente: {str(e)}"

        elif accion == "modificar":
            cliente_id = int(request.form.get('id'))
            cliente = Cliente.query.get(cliente_id)
            if cliente:
                cliente.cedula = request.form.get('cedula').strip()
                cliente.nombres = request.form.get('nombre').strip()
                cliente.direccion = request.form.get('direccion').strip()
                fecha_inst = request.form.get('fecha_instalacion')
                cliente.fecha_instalacion = datetime.strptime(fecha_inst, "%Y-%m-%d") if fecha_inst else None
                cliente.ip = request.form.get('ip').strip()
                try:
                    db.session.commit()
                    mensaje = f"✅ Cliente {cliente_id} modificado con éxito."
                except Exception as e:
                    db.session.rollback()
                    mensaje = f"❌ Error al modificar cliente: {str(e)}"

        elif accion == "eliminar":
            cliente_id = int(request.form.get('id'))
            cliente = Cliente.query.get(cliente_id)
            if cliente:
                try:
                    db.session.delete(cliente)
                    db.session.commit()
                    mensaje = f"❌ Cliente {cliente_id} eliminado."
                except Exception as e:
                    db.session.rollback()
                    mensaje = f"❌ Error al eliminar cliente: {str(e)}"

        elif accion == "buscar":
            termino = request.form.get('buscar').strip().lower()
            clientes = Cliente.query.filter(
                (Cliente.nombres.ilike(f"%{termino}%")) |
                (Cliente.ip.ilike(f"%{termino}%"))
            ).all()
            return render_template('registro_cliente.html', clientes=clientes, mensaje=mensaje)

    clientes = Cliente.query.all()
    return render_template('registro_cliente.html', clientes=clientes, mensaje=mensaje)

# -------------------- Context Processor --------------------
@app.context_processor
def inject_user_role():
    return dict(user_role=session.get('role'))

# -------------------- Logout --------------------
@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

# -------------------- Run --------------------
if __name__ == '__main__':
    app.run(debug=True)
